import sys
import os
from openpyxl import Workbook


PROD_DICT = {}
SIT_DICT = {}

#TODO: create gitignore
#TODO: output to excel file

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def read_file(input, dict):

    INPUT_FILE = open(input, 'r')
    INPUT_FILE_LINES = INPUT_FILE.readlines()

    # Strips the newline character
    count = 1
    INPUT_OBJECT_LIST = []

    start_reading = False
    RIC = ""
    DOMAIN = ""

    for line in INPUT_FILE_LINES:

        # print("Prod-{}: {}".format(count, line.strip()))
        # count = count + 1

        # < it: RIC > DXVF21 < / it: RIC >
        # < it: DOMAIN > MARKET_BY_PRICE < / it: DOMAIN >

        if line.find('<exlObject>') != -1 and start_reading != True:
            # print('start reading exlObject')
            start_reading = True
            # print("Prod-StartReading{}: {}".format(count, line.strip()))
            continue

        if start_reading:
            INPUT_OBJECT_LIST.append(line)

        RIC_index_start = line.find('<it:RIC>')
        RIC_index_end = line.find('</it:RIC>')

        DOMAIN_index_start = line.find('<it:DOMAIN>')
        DOMAIN_index_end = line.find('</it:DOMAIN>')

        # print(str(RIC_index_start) + ' ' + str(RIC_index_end))
        if RIC_index_start != -1 and RIC_index_end != -1:
            RIC = line[RIC_index_start + 8:RIC_index_end]

        if DOMAIN_index_start != -1 and DOMAIN_index_end != -1:
            DOMAIN = line[DOMAIN_index_start + 11:DOMAIN_index_end]

        if line.find('</exlObject>') != -1:
            # add to dict
            KEY = RIC + "-" + DOMAIN
            # print(KEY)
            # print(PROD_OBJECT_LIST)
            # print("Prod-EndReading{}: {}".format(count, line.strip()))
            dict[KEY] = INPUT_OBJECT_LIST
            start_reading = False
            RIC = ""
            DOMAIN = ""
            # print(dict)
            INPUT_OBJECT_LIST = []
    # print(dict)

def compare_files(input_prod, input_sit, sheet, filename):

    # './PROD/BUSE_DXV_MBP_OMM.EXL.PROD.txt'
    # SIT_FILE = open('./SIT/BUSE_DXV_MBP_OMM.EXL.SIT', 'r')
    # SIT_FILE_LINES = SIT_FILE.readlines()
    read_file(input_prod, PROD_DICT)
    read_file(input_sit, SIT_DICT)

    # for elem in PROD_DICT:
    #     print(PROD_DICT[elem])

    # print('-------- BAZINGA --------------')

    # for elem in SIT_DICT:
    #     print(SIT_DICT[elem])

    # quantity of RICs in file
    # print(len(PROD_DICT))
    # print(len(SIT_DICT))

    if PROD_DICT != SIT_DICT:
        PROD_KEYS = PROD_DICT.keys()
        SIT_KEYS = SIT_DICT.keys()
        # print(keys)

        COMP_FILE_DIFF_UNIQUE = []

        # 1. PROD vs SIT
        PROD_DIFF = []
        ABSENT_PROD_KEYS_ON_SIT = []
        for key in PROD_KEYS:
            if key in SIT_DICT:
                if PROD_DICT[key] != SIT_DICT[key]:
                    PROD_DIFF.append(key)
                    if key not in COMP_FILE_DIFF_UNIQUE:
                        COMP_FILE_DIFF_UNIQUE.append(key)
                    # print("AAAAAAAAAAAAAAAAA    --------- " + key)
            else:
                ABSENT_PROD_KEYS_ON_SIT.append(key)
                # print('No such key -> ' + key + ' in file -> ' + input_prod )

        # 2. SIT vs PROD
        SIT_DIFF = []
        ABSENT_SIT_KEYS_ON_PROD = []
        for key in SIT_KEYS:
            if key in PROD_DICT:
                if SIT_DICT[key] != PROD_DICT[key]:
                    SIT_DIFF.append(key)
                    if key not in COMP_FILE_DIFF_UNIQUE:
                        COMP_FILE_DIFF_UNIQUE.append(key)
                    # print("BBBBBBBBBBBBBBBBBB    --------- " + key)
            else:
                ABSENT_SIT_KEYS_ON_PROD.append(key)
                # print('No such key -> ' + key + ' in file -> ' + input_prod)

        print("ABSENT_PROD_KEYS_ON_SIT -> " + str(ABSENT_PROD_KEYS_ON_SIT))
        print("ABSENT_SIT_KEYS_ON_PROD -> " + str(ABSENT_SIT_KEYS_ON_PROD))
        print("COMP_FILE_DIFF_UNIQUE -> " + str(COMP_FILE_DIFF_UNIQUE))

        i = 2
        for elem in ABSENT_PROD_KEYS_ON_SIT:
            sheet.cell(row=i, column=1).value = elem + '/' + filename
            i = i + 1

        i=2
        for elem in ABSENT_SIT_KEYS_ON_PROD:
            sheet.cell(row=i, column=2).value = elem + '/' +  filename
            i = i + 1

        i=2
        for elem in COMP_FILE_DIFF_UNIQUE:
            sheet.cell(row=i, column=3).value = elem + '/' +  filename
            i = i + 1
    else:
        print("equal")

    PROD_DICT.clear()
    SIT_DICT.clear()

def move_files(PROD_DIR, SIT_DIR):
    subfolders_prod = [f.path for f in os.scandir(PROD_DIR) if f.is_dir()]

    print(subfolders_prod)
    for subfolder in subfolders_prod:
        print(os.listdir(subfolder))

    subfolders_sit = [f.path for f in os.scandir(SIT_DIR) if f.is_dir()]
    print(subfolders_sit)
    for subfolder in subfolders_sit:
        print(os.listdir(subfolder))

if __name__ == '__main__':
    workbook = Workbook()
    sheet = workbook.active

    PROD_DIR = sys.argv[1]
    SIT_DIR = sys.argv[2]

    move_files(PROD_DIR, SIT_DIR)

    PROD_DIR_FILES = os.listdir(PROD_DIR)
    SIT_DIR_FILES = os.listdir(SIT_DIR)

    COMMON_FILES = []
    UNCOMMON_FILES = []

    for file in SIT_DIR_FILES:
        if file in PROD_DIR_FILES:
            COMMON_FILES.append(file)
        else:
            UNCOMMON_FILES.append(file)

    sheet.cell(row=1, column=1).value = 'ABSENT_PROD_KEYS_ON_SIT'
    sheet.cell(row=1, column=2).value = 'ABSENT_SIT_KEYS_ON_PROD'
    sheet.cell(row=1, column=3).value = 'COMP_FILE_DIFF_UNIQUE'

    for file in COMMON_FILES:
        print(f"{bcolors.HEADER}file analyzed -> {file}{bcolors.ENDC}")
        compare_files(PROD_DIR + file, SIT_DIR + file, sheet, file)

    print(f"\n{bcolors.OKGREEN}Analyzed files:\n{COMMON_FILES}{bcolors.ENDC}")
    print(f"\n{bcolors.WARNING}UNCOMMON_FILES. NOT ANALIZED:\n{UNCOMMON_FILES}{bcolors.ENDC}")

    # print("ABSENT_PROD_KEYS_ON_SIT -> " + str(ABSENT_PROD_KEYS_ON_SIT))
    # print("ABSENT_SIT_KEYS_ON_PROD -> " + str(ABSENT_SIT_KEYS_ON_PROD))
    # print("COMP_FILE_DIFF_UNIQUE -> " + str(COMP_FILE_DIFF_UNIQUE))

    workbook.save(filename="DIFF_RESULT.xlsx")


